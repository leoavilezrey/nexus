#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Genera el plan de pruebas manual en formato Excel (.xlsx) con formato amigable."""

import sys, os
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

NEXUS_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if NEXUS_ROOT not in sys.path:
    sys.path.insert(0, NEXUS_ROOT)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Plan de Pruebas"

# ── Estilos ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
SUBSEC_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
SUBSEC_FONT = Font(name="Calibri", size=10, bold=True, color="2962A7")
NORMAL_FONT = Font(name="Calibri", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ── Columnas ──
COLS = ["#", "Sección", "Sub-Menú", "Acción a Realizar", "¿Funciona?\n(Sí/No)", "Mensaje de Error", "Ubicación en Código", "Comentarios"]
COL_WIDTHS = [8, 18, 22, 50, 14, 30, 28, 30]

# ── Título ──
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = "📋 Plan de Pruebas Manual Exhaustivo — Nexus Dashboard"
title_cell.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:H2')
ws['A2'].value = "Fecha: ___/___/2026    Tester: _________________    Versión: Post-Corrección v3"
ws['A2'].font = Font(name="Calibri", size=10, italic=True, color="666666")
ws['A2'].alignment = Alignment(horizontal='center')

# ── Cabecera ──
for col_idx, (header, width) in enumerate(zip(COLS, COL_WIDTHS), 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[4].height = 30

# ── Datos de prueba ──
# Formato: (num, seccion, submenu, accion, ubicacion)  | las columnas 5,6,8 se dejan vacías

tests = [
    # ── PASO 0: ARRANQUE ──
    ("SEC", "0. ARRANQUE DE LA APLICACIÓN", "", "", ""),
    ("0.1", "0. Arranque", "", "Ejecutar .\\venv\\Scripts\\python.exe main.py", "main.py:L1-10"),
    ("0.2", "0. Arranque", "", "Banner NEXUS visible al arrancar", "dashboard.py:L85-93"),
    ("0.3", "0. Arranque", "", "Panel de estadísticas (Registros/Cards/Tags/Links)", "dashboard.py:L95-121"),
    ("0.4", "0. Arranque", "", "Grilla de 6 paneles del menú principal", "dashboard.py:L1754-1787"),
    ("0.5", "0. Arranque", "", "Barra de comandos rápidos visible", "dashboard.py:L1788-1789"),
    ("0.6", "0. Arranque", "", "Console theme carga sin errores", "dashboard.py:L5-12"),

    # ── MENÚ 1: AGREGAR ──
    ("SEC", "1. MENÚ AGREGAR — Captura de Contenido", "", "", ""),
    ("1.0", "1. Agregar", "", "Presionar 1 → entra al submenú Agregar", "dashboard.py:L128-164"),
    ("1.1", "1. Agregar", "", "Muestra opciones (1,2,3,4,6,S,0)", "dashboard.py:L153-159"),
    ("1.2", "1. Agregar", "", "Destino actual dice 'CORE (Local SSD)'", "dashboard.py:L139-141"),
    
    # 1.1 Archivo Local
    ("SUB", "", "1.1 Añadir Archivo Local", "", ""),
    ("1.1.1", "1. Agregar", "1.1 Archivo Local", "Presionar 1 → solicita ruta del archivo", "dashboard.py:L183"),
    ("1.1.2", "1. Agregar", "1.1 Archivo Local", "Presionar Enter SIN escribir → '⚠ Ruta vacía. Operación cancelada.'", "dashboard.py:L184-187"),
    ("1.1.3", "1. Agregar", "1.1 Archivo Local", "Ingresar ruta válida (ej. main.py)", "file_manager.py:L22-90"),
    ("1.1.4", "1. Agregar", "1.1 Archivo Local", "Ingresar tags: test, prueba", "dashboard.py:L188"),
    ("1.1.5", "1. Agregar", "1.1 Archivo Local", "Muestra '✅ Archivo indexado correctamente'", "dashboard.py:L199"),
    ("1.1.6", "1. Agregar", "1.1 Archivo Local", "Panel de detalles: ID, Tipo, Título, Ubicación", "dashboard.py:L200-215"),
    ("1.1.7", "1. Agregar", "1.1 Archivo Local", "Vista previa del contenido visible", "dashboard.py:L215-221"),
    ("1.1.8", "1. Agregar", "1.1 Archivo Local", "content_raw = None → no crash (guard activo)", "dashboard.py:L215 (BUG FIJADO)"),
    ("1.1.9", "1. Agregar", "1.1 Archivo Local", "Preguntan editar → responder 'n'", "dashboard.py:L223"),
    ("1.1.10", "1. Agregar", "1.1 Archivo Local", "Navegación: 1=Añadir otro, 2=Volver, 0=Principal", "dashboard.py:L224-228"),
    ("1.1.11", "1. Agregar", "1.1 Archivo Local", "Repetir con ruta INVÁLIDA → error claro, sin crash", "file_manager.py:L24"),
    ("1.1.12", "1. Agregar", "1.1 Archivo Local", "⚡ Ctrl+C mientras espera ruta → debe volver al menú SIN crash", "main.py:L1821-1825"),
    ("1.1.13", "1. Agregar", "1.1 Archivo Local", "⚡ Ctrl+C durante descarga → debe abortar limpiamente", "dashboard.py:L192-193"),

    # 1.2 URL
    ("SUB", "", "1.2 Añadir URL (Web/YouTube)", "", ""),
    ("1.2.1", "1. Agregar", "1.2 URL Web/YouTube", "Presionar 2 → solicita URL", "dashboard.py:L236"),
    ("1.2.2", "1. Agregar", "1.2 URL Web/YouTube", "Presionar Enter SIN escribir → '⚠ URL vacía'", "dashboard.py:L237-240"),
    ("1.2.3", "1. Agregar", "1.2 URL Web/YouTube", "URL genérica (ej. https://example.com)", "web_scraper.py:L198-252"),
    ("1.2.4", "1. Agregar", "1.2 URL Web/YouTube", "Spinner 'Descargando...' visible", "dashboard.py:L244"),
    ("1.2.5", "1. Agregar", "1.2 URL Web/YouTube", "Resultado ✅ o mensaje de fallo claro", "dashboard.py:L247-265"),
    ("1.2.6", "1. Agregar", "1.2 URL Web/YouTube", "URL contenido None → '⚠ No se pudo extraer' (guard)", "dashboard.py:L263-272 (BUG FIJADO)"),
    ("1.2.7", "1. Agregar", "1.2 URL Web/YouTube", "Probar con URL de YouTube (video corto)", "web_scraper.py:L49-141"),
    ("1.2.8", "1. Agregar", "1.2 URL Web/YouTube", "Transcripción extraída correctamente", "web_scraper.py:L94-121"),
    ("1.2.9", "1. Agregar", "1.2 URL Web/YouTube", "URL inválida/malformada → error controlado", "web_scraper.py:L207-218"),

    # 1.3 Nota
    ("SUB", "", "1.3 Escribir Nota Libre", "", ""),
    ("1.3.1", "1. Agregar", "1.3 Nota Libre", "Presionar 3 → solicita título", "dashboard.py:L289"),
    ("1.3.2", "1. Agregar", "1.3 Nota Libre", "Título: 'Nota de Prueba'", "dashboard.py:L290"),
    ("1.3.3", "1. Agregar", "1.3 Nota Libre", "Se abre Bloc de notas (notepad)", "pkm_manager.py:L27"),
    ("1.3.4", "1. Agregar", "1.3 Nota Libre", "Escribir algo, guardar y cerrar → '✅ Nota almacenada'", "dashboard.py:L310-319"),
    ("1.3.5", "1. Agregar", "1.3 Nota Libre", "Cerrar bloc SIN escribir → 'Nota vacía. Abortando'", "dashboard.py:L314-316 (BUG FIJADO)"),

    # 1.4 App
    ("SUB", "", "1.4 Aplicación/Herramienta", "", ""),
    ("1.4.1", "1. Agregar", "1.4 App/Herramienta", "Presionar 4 → formulario completo", "dashboard.py:L345-399"),
    ("1.4.2", "1. Agregar", "1.4 App/Herramienta", "Llenar todos los campos → '✅ App registrada'", "dashboard.py:L393"),
    ("1.4.3", "1. Agregar", "1.4 App/Herramienta", "Sin crash por nx_db shadowing (BUG FIJADO)", "dashboard.py:L353"),

    # 1.6 Pipeline
    ("SUB", "", "1.6 Pipeline YouTube", "", ""),
    ("1.6.1", "1. Agregar", "1.6 Pipeline", "Presionar 6 → inicia pipeline sin crash", "pipeline_manager.py:L30-144"),

    # 1.S Destino
    ("SUB", "", "1.S Cambiar Destino", "", ""),
    ("1.S.1", "1. Agregar", "1.S Destino", "S → 'STAGING Activado' (si G: montado)", "dashboard.py:L170-177"),
    ("1.S.2", "1. Agregar", "1.S Destino", "S otra vez → 'CORE reactivado'", "dashboard.py:L174"),
    ("1.0.1", "1. Agregar", "Volver", "Presionar 0 → menú principal", "dashboard.py:L163-164"),

    # ── MENÚ 2: GESTIONAR (DETALLADO) ──
    ("SEC", "2. MENÚ GESTIONAR — Explorador Maestro (DETALLADO)", "", "", ""),
    ("2.0", "2. Gestionar", "", "Presionar 2 → entra al Explorador", "dashboard.py:L407-643"),
    ("2.1", "2. Gestionar", "", "Tabla visible: ID, Tipo, Título, Ext, Tags, Resumen, FC", "dashboard.py:L471-502"),
    ("2.2", "2. Gestionar", "", "Columnas correctamente alineadas", "dashboard.py:L474-481"),
    ("2.3", "2. Gestionar", "", "IDs visibles y legibles", "dashboard.py:L496"),

    # 2.P Paginación
    ("SUB", "", "2.P Paginación", "", ""),
    ("2.P.1", "2. Gestionar", "2.P Paginación", "→ (flecha derecha) → siguiente página", "dashboard.py:L516-519"),
    ("2.P.2", "2. Gestionar", "2.P Paginación", "En última página → '⚠ Ya estás en la última página'", "dashboard.py:L518-519"),
    ("2.P.3", "2. Gestionar", "2.P Paginación", "← (flecha izquierda) → página anterior", "dashboard.py:L520-521"),
    ("2.P.4", "2. Gestionar", "2.P Paginación", "En primera página, presionar ← → se mantiene sin error", "dashboard.py:L520"),
    ("2.P.5", "2. Gestionar", "2.P Paginación", "Número de página visible en título de tabla", "dashboard.py:L471"),

    # 2.F Filtros
    ("SUB", "", "2.F Filtros del Explorador", "", ""),
    ("2.F.1", "2. Gestionar", "2.F Filtros", "Q → abre panel de filtros inteligente", "dashboard.py:L525-534"),
    ("2.F.2", "2. Gestionar", "2.F Filtros", "Escribir 'python' → filtra por nombre", "search_engine.py:L36-43"),
    ("2.F.3", "2. Gestionar", "2.F Filtros", "Tabla muestra solo registros que contienen 'python'", "dashboard.py:L440-465"),
    ("2.F.4", "2. Gestionar", "2.F Filtros", "L → limpiar todos los filtros", "dashboard.py:L522-524"),
    ("2.F.5", "2. Gestionar", "2.F Filtros", "Resultado regresa a vista completa (sin filtros)", "dashboard.py:L523"),
    ("2.F.6", "2. Gestionar", "2.F Filtros", "Q → escribir 't:YouTube_Pipeline' → filtra por tag", "search_engine.py:L58-61"),
    ("2.F.7", "2. Gestionar", "2.F Filtros", "Tabla muestra solo registros con ese tag", "dashboard.py:L440-465"),
    ("2.F.8", "2. Gestionar", "2.F Filtros", "L → limpiar", "dashboard.py:L522-524"),
    ("2.F.9", "2. Gestionar", "2.F Filtros", "Q → 'e:pdf' → filtra por extensión", "search_engine.py:L68-95"),
    ("2.F.10", "2. Gestionar", "2.F Filtros", "Q → 'e:web' → filtra tipo web", "search_engine.py:L68-95"),
    ("2.F.11", "2. Gestionar", "2.F Filtros", "Q → 'e:youtube' → filtra tipo youtube", "search_engine.py:L68-95"),
    ("2.F.12", "2. Gestionar", "2.F Filtros", "L → limpiar", "dashboard.py:L522-524"),
    ("2.F.13", "2. Gestionar", "2.F Filtros", "Q → 'i:1-10' → muestra solo IDs 1 a 10", "search_engine.py:L143-159"),
    ("2.F.14", "2. Gestionar", "2.F Filtros", "Q → 'i:5,10,15' → muestra solo esos 3 IDs", "search_engine.py:L143-159"),
    ("2.F.15", "2. Gestionar", "2.F Filtros", "Q → 'h:y' → registros con has_info", "search_engine.py:L96-115"),
    ("2.F.16", "2. Gestionar", "2.F Filtros", "Q → 's:y' → registros que son fuente de flashcards", "search_engine.py:L162-176"),
    ("2.F.17", "2. Gestionar", "2.F Filtros", "Q → filtro vacío (Enter) → limpiar filtros", "dashboard.py:L530-534"),
    ("2.F.18", "2. Gestionar", "2.F Filtros", "Combinación: 'python t:docs e:pdf' → múltiples filtros simultáneos", "search_engine.py:L204-254"),
    ("2.F.19", "2. Gestionar", "2.F Filtros", "Indicador 'Filtros: Sí' visible en título de tabla", "dashboard.py:L471"),

    # 2.D Vista Detallada
    ("SUB", "", "2.D Vista Detallada de Registro", "", ""),
    ("2.D.1", "2. Gestionar", "2.D Detalle", "Escribir un ID visible → Enter → panel detallado", "dashboard.py:L536-542"),
    ("2.D.2", "2. Gestionar", "2.D Detalle", "Título, tipo, ruta, contenido, tags, resumen → todos visibles", "dashboard.py:L645-750"),
    ("2.D.3", "2. Gestionar", "2.D Detalle", "ID inexistente → mensaje de error claro", "dashboard.py:L540-541"),
    ("2.D.4", "2. Gestionar", "2.D Detalle", "e → editar campo → seleccionar campo → funciona", "dashboard.py:L760-840"),
    ("2.D.5", "2. Gestionar", "2.D Detalle", "sum → generar resumen IA (o fallback sin API)", "summary_agent.py:L19-69"),
    ("2.D.6", "2. Gestionar", "2.D Detalle", "fc → generar flashcards IA (o mockup)", "study_agent.py:L27-147"),
    ("2.D.7", "2. Gestionar", "2.D Detalle", "fc+ → crear flashcard manual → ingresar Q y A", "dashboard.py:L870-920"),
    ("2.D.8", "2. Gestionar", "2.D Detalle", "'Flashcard manual creada' confirmado", "dashboard.py:L910"),
    ("2.D.9", "2. Gestionar", "2.D Detalle", "ver fc → tabla de flashcards del registro", "dashboard.py:L930-960"),
    ("2.D.10", "2. Gestionar", "2.D Detalle", "links → ver conexiones neuronales", "dashboard.py:L960-990"),
    ("2.D.11", "2. Gestionar", "2.D Detalle", "abrir → abre archivo/URL original en sistema", "study_engine.py:L61-119"),
    ("2.D.12", "2. Gestionar", "2.D Detalle", "mut → mutar flashcards difíciles", "mutation_agent.py"),
    ("2.D.13", "2. Gestionar", "2.D Detalle", "p → volver a la lista del explorador", "dashboard.py:L543"),

    # 2.X Eliminar
    ("SUB", "", "2.X Eliminar Registros", "", ""),
    ("2.X.1", "2. Gestionar", "2.X Eliminar", "del [ID] → solicita confirmación", "dashboard.py:L545-570"),
    ("2.X.2", "2. Gestionar", "2.X Eliminar", "Confirmar → registro eliminado correctamente", "dashboard.py:L560-565"),
    ("2.X.3", "2. Gestionar", "2.X Eliminar", "del [ID1,ID2,ID3] → eliminación en lote", "dashboard.py:L545-570"),
    ("2.X.4", "2. Gestionar", "2.X Eliminar", "Cancelar eliminación → registro preservado", "dashboard.py:L555"),

    # 2.V Vincular
    ("SUB", "", "2.V Vincular Registros", "", ""),
    ("2.V.1", "2. Gestionar", "2.V Vincular", "m [ID1] [ID2] → vincular manualmente", "dashboard.py:L575-590"),
    ("2.V.2", "2. Gestionar", "2.V Vincular", "ia [ID1] [ID2] → vincular con IA", "dashboard.py:L590-610"),
    ("2.V.3", "2. Gestionar", "2.V Vincular", "0 → volver al menú principal", "dashboard.py:L514-515"),

    # ── MENÚ 3: ACTIVE RECALL ──
    ("SEC", "3. MENÚ ACTIVE RECALL — Pomodoro SRS", "", "", ""),
    ("3.0", "3. Active Recall", "", "Presionar 3 → entra a Active Recall", "dashboard.py:L1091-1104"),
    ("3.1", "3. Active Recall", "", "Panel 'Motor Pomodoro Listo' + cards pendientes", "dashboard.py:L1109-1121"),
    ("3.2", "3. Active Recall", "", "Tabla fuentes: ID, Título, Tags, Pend/Tot, Flashcard", "dashboard.py:L1149-1195"),
    ("3.N.1", "3. Active Recall", "3.N Navegación", "→ siguiente, ← anterior, Q filtrar, L limpiar", "dashboard.py:L1212-1231"),
    ("3.P.1", "3. Active Recall", "3.P Pomodoro", "pm → sesión Pomodoro 25min → pregunta visible", "study_engine.py:L146-435"),
    ("3.P.2", "3. Active Recall", "3.P Pomodoro", "Enter → respuesta → calificar (1-4)", "study_engine.py:L240-380"),
    ("3.P.3", "3. Active Recall", "3.P Pomodoro", "q o 0 → terminar sesión antes", "study_engine.py:L390"),
    ("3.P.4", "3. Active Recall", "3.P Pomodoro", "Resumen final de sesión al terminar", "study_engine.py:L395-430"),
    ("3.A.1", "3. Active Recall", "3.A Adelantar", "pa → adelantar repasos → funciona", "dashboard.py:L1001-1089"),
    ("3.G.1", "3. Active Recall", "3.G Flashcards", "ia [ID] → generar con IA", "dashboard.py:L1250-1320"),
    ("3.G.2", "3. Active Recall", "3.G Flashcards", "man [ID] → crear manual", "dashboard.py:L1320-1370"),
    ("3.D.1", "3. Active Recall", "3.D Eliminar", "del [IDs] → eliminar flashcards", "dashboard.py:L1380-1420"),
    ("3.0.1", "3. Active Recall", "Volver", "0 → menú principal", "dashboard.py:L1210-1211"),

    # ── MENÚ 4: ESTADÍSTICAS ──
    ("SEC", "4. MENÚ ESTADÍSTICAS", "", "", ""),
    ("4.0", "4. Estadísticas", "", "Presionar 4 → entra", "dashboard.py:L1655-1660"),
    ("4.1", "4. Estadísticas", "", "Panel 4.1: Composición del Cerebro (conteo por tipo)", "dashboard.py:L1672-1680"),
    ("4.2", "4. Estadísticas", "", "Panel 4.2: Red Neuronal (vínculos, tags)", "dashboard.py:L1682-1687"),
    ("4.3", "4. Estadísticas", "", "Panel 4.3: Madurez Cognitiva SRS", "dashboard.py:L1689-1698"),
    ("4.G.1", "4. Estadísticas", "4.G Drive", "G → exportar a Google Drive", "dashboard.py:L1714-1728"),
    ("4.S.1", "4. Estadísticas", "4.S Filtrar", "S → filtrar por tag/tipo", "dashboard.py:L1730-1735"),
    ("4.0.1", "4. Estadísticas", "Volver", "Enter → menú principal", "dashboard.py:L1737-1738"),

    # ── MENÚ 5: SALIR ──
    ("SEC", "5. SALIR", "", "", ""),
    ("5.1", "5. Salir", "", "Escribir 5 → 'Cerrando módulos de Nexus'", "dashboard.py:L1801-1804"),
    ("5.2", "5. Salir", "", "Terminal libre, sin errores ni tracebacks", "main.py"),

    # ── OMNIBAR (DETALLADO) ──
    ("SEC", "6. OMNIBAR — Búsqueda Directa desde Menú Principal (DETALLADO)", "", "", ""),
    ("6.1", "6. Omnibar", "", "Escribir texto libre (ej. 'python') → omnibar se activa", "dashboard.py:L1805-1812"),
    ("6.2", "6. Omnibar", "", "Mensaje: 'Omnibar: Saltando al Gestor con python'", "dashboard.py:L1810"),
    ("6.3", "6. Omnibar", "", "Se abre menú Gestionar con filtro aplicado", "dashboard.py:L1812"),
    ("6.4", "6. Omnibar", "", "Resultados filtrados corresponden al término buscado", "dashboard.py:L407-465"),
    ("6.5", "6. Omnibar", "", "Paginación funciona dentro de la búsqueda omnibar", "dashboard.py:L516-521"),
    ("6.6", "6. Omnibar", "", "Presionar 0 desde gestor → volver al menú principal", "dashboard.py:L514-515"),
    ("6.7", "6. Omnibar", "", "Escribir ':comando_invalido' → 'Comando desconocido' en amarillo", "dashboard.py:L1806-1809"),
    ("6.8", "6. Omnibar", "", "':' vacío (solo dos puntos) → 'Comando desconocido: :'", "dashboard.py:L1806"),
    ("6.9", "6. Omnibar", "", "Escribir texto con espacios → búsqueda funciona", "dashboard.py:L1810"),
    ("6.10", "6. Omnibar", "", "Escribir solo espacios → NO activa omnibar (strip)", "dashboard.py:L1790, L1805"),
    ("6.11", "6. Omnibar", "", "Omnibar con término que no existe → tabla vacía, sin crash", "dashboard.py:L1812 + search"),
    ("6.12", "6. Omnibar", "", "Escribir 'q', 'exit', 'quit' → cierra Nexus", "dashboard.py:L1801"),
    ("6.13", "6. Omnibar", "", "Escribir '0' → cierra Nexus", "dashboard.py:L1801"),

    # ── CTRL+C (DETALLADO) ──
    ("SEC", "7. CTRL+C — Interrupción Segura (DETALLADO)", "", "", ""),
    ("7.1", "7. Ctrl+C", "", "Ctrl+C en menú principal → 'Interrupción detectada. Saliendo...'", "main.py:L1821-1825"),
    ("7.2", "7. Ctrl+C", "", "Terminal limpia, sin traceback de Python", "main.py:L1823-1825"),
    ("7.3", "7. Ctrl+C", "", "Ctrl+C en Menú 1 (Agregar) → sale limpiamente al menú o cierra", "dashboard.py:L135 + main_loop handler"),
    ("7.4", "7. Ctrl+C", "", "Ctrl+C mientras espera ruta archivo (opción 1.1) → no crash", "dashboard.py:L183 (Prompt.ask)"),
    ("7.5", "7. Ctrl+C", "", "Ctrl+C mientras espera URL (opción 1.2) → no crash", "dashboard.py:L236 (Prompt.ask)"),
    ("7.6", "7. Ctrl+C", "", "Ctrl+C en Menú 2 (Gestionar) → sale limpiamente", "dashboard.py:L407"),
    ("7.7", "7. Ctrl+C", "", "Ctrl+C en Menú 3 (Active Recall) → sale limpiamente", "dashboard.py:L1091"),
    ("7.8", "7. Ctrl+C", "", "Ctrl+C durante sesión Pomodoro → sale sin corromper DB", "study_engine.py:L146"),
    ("7.9", "7. Ctrl+C", "", "Ctrl+C en Menú 4 (Estadísticas) → sale limpiamente", "dashboard.py:L1655"),
    ("7.10", "7. Ctrl+C", "", "Después de Ctrl+C, relanzar main.py → funciona normalmente", "main.py + nexus.db"),

    # ── LIMPIEZA ──
    ("SEC", "8. LIMPIEZA POST-PRUEBA", "", "", ""),
    ("8.1", "8. Limpieza", "", "Menú 2 → buscar registros creados durante la prueba", "dashboard.py:L407"),
    ("8.2", "8. Limpieza", "", "del [IDs de prueba] → eliminar + confirmar", "dashboard.py:L545-570"),
    ("8.3", "8. Limpieza", "", "Verificar que ya no aparecen en búsqueda", "search_engine.py"),
]

# ── Escribir datos ──
row = 5
for item in tests:
    num, sec, sub, act, loc = item
    
    if num == "SEC":
        # Fila de sección
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value=sec)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 25
    elif num == "SUB":
        # Fila de sub-sección
        ws.merge_cells(f'A{row}:H{row}')
        cell = ws.cell(row=row, column=1, value=f"    ▸ {sub}")
        cell.font = SUBSEC_FONT
        cell.fill = SUBSEC_FILL
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 22
    else:
        # Fila de test
        values = [num, sec, sub, act, "", "", loc, ""]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx in [4, 6, 7, 8]))
            cell.border = THIN_BORDER
        # Columna "¿Funciona?" con validación visual
        func_cell = ws.cell(row=row, column=5)
        func_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1

# ── Resumen al final ──
row += 1
ws.merge_cells(f'A{row}:H{row}')
ws.cell(row=row, column=1, value="RESUMEN GENERAL").font = SECTION_FONT
ws.cell(row=row, column=1).fill = SECTION_FILL
row += 1

summary_data = [
    ("0. Arranque", 6), ("1. Agregar", 30), ("2. Gestionar", 30),
    ("3. Active Recall", 14), ("4. Estadísticas", 7), ("5. Salir", 2),
    ("6. Omnibar", 13), ("7. Ctrl+C", 10), ("8. Limpieza", 3)
]

for s_name, s_count in summary_data:
    ws.cell(row=row, column=1, value=s_name).font = NORMAL_FONT
    ws.cell(row=row, column=2, value=s_count).font = NORMAL_FONT
    ws.cell(row=row, column=3, value="").font = NORMAL_FONT  # Pasaron
    ws.cell(row=row, column=4, value="").font = NORMAL_FONT  # Fallaron
    for c in range(1, 5):
        ws.cell(row=row, column=c).border = THIN_BORDER
    row += 1

ws.cell(row=row, column=1, value="TOTAL").font = Font(name="Calibri", size=11, bold=True)
ws.cell(row=row, column=2, value=sum(c for _, c in summary_data)).font = Font(name="Calibri", size=11, bold=True)
for c in range(1, 5):
    ws.cell(row=row, column=c).border = THIN_BORDER

# ── Guardar ──
output_nexus = os.path.join(NEXUS_ROOT, "docs", "plan_pruebas_manual_nexus.xlsx")
wb.save(output_nexus)
print(f"Excel generado: {output_nexus}")

# Intentar guardar también en OneDrive
onedrive_paths = [
    os.path.expandvars(r"%USERPROFILE%\OneDrive\Documents"),
    os.path.expandvars(r"%USERPROFILE%\OneDrive - Personal\Documents"),
    r"C:\Users\DELL\OneDrive\Documents",
]

for od_path in onedrive_paths:
    if os.path.isdir(od_path):
        od_file = os.path.join(od_path, "plan_pruebas_manual_nexus1.xlsx")
        try:
            wb.save(od_file)
            print(f"Copia en OneDrive: {od_file}")
        except Exception as e:
            print(f"No se pudo guardar en OneDrive: {e}")
        break
else:
    print("OneDrive path no encontrado — archivo guardado solo en docs/")

print(f"Total items de prueba: {sum(c for _, c in summary_data)}")
